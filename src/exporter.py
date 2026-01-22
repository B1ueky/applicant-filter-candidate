"""Excel exporter for candidate data."""

import os
from typing import List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .candidate import Candidate


class ExcelExporter:
    """
    Export filtered candidates to Excel format.

    Output columns: Name | Position | Experience | Location | LinkedIn URL
    """

    def __init__(self, output_directory: str = "output"):
        """
        Initialize the exporter.

        Args:
            output_directory: Directory to save output files
        """
        self.output_directory = output_directory
        self._ensure_output_directory()

    def _ensure_output_directory(self) -> None:
        """Create output directory if it doesn't exist."""
        if not os.path.exists(self.output_directory):
            os.makedirs(self.output_directory)

    def export(
        self,
        candidates: List[Candidate],
        filename: str = "candidates.xlsx"
    ) -> str:
        """
        Export candidates to Excel file.

        Args:
            candidates: List of candidates to export
            filename: Output filename

        Returns:
            Full path to the exported file
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Filtered Candidates"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Headers
        headers = ["Name", "Position", "Experience (Years)", "Location", "LinkedIn URL"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row, candidate in enumerate(candidates, 2):
            data = [
                candidate.name,
                candidate.current_position,
                candidate.experience_years,
                candidate.location,
                candidate.linkedin_url
            ]
            for col, value in enumerate(data, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col == 5:  # LinkedIn URL column
                    cell.alignment = Alignment(horizontal="left")

        # Adjust column widths
        column_widths = [25, 30, 18, 15, 45]
        for col, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(col)].width = width

        # Freeze header row
        sheet.freeze_panes = "A2"

        # Save file
        filepath = os.path.join(self.output_directory, filename)
        workbook.save(filepath)

        return filepath

    def export_detailed(
        self,
        candidates: List[Candidate],
        filename: str = "candidates_detailed.xlsx"
    ) -> str:
        """
        Export candidates with detailed information.

        Includes all candidate fields.

        Args:
            candidates: List of candidates to export
            filename: Output filename

        Returns:
            Full path to the exported file
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Detailed Candidates"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Headers
        headers = [
            "Name", "Age", "Position", "Experience (Years)",
            "Location", "Nationality", "Education Background",
            "Work Background", "Skills", "Languages", "LinkedIn URL"
        ]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Data rows
        for row, candidate in enumerate(candidates, 2):
            data = [
                candidate.name,
                candidate.age,
                candidate.current_position,
                candidate.experience_years,
                candidate.location,
                candidate.nationality,
                ", ".join(candidate.education_background),
                ", ".join(candidate.work_background),
                ", ".join(candidate.skills),
                ", ".join(candidate.languages),
                candidate.linkedin_url
            ]
            for col, value in enumerate(data, 1):
                sheet.cell(row=row, column=col, value=value)

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 20

        # LinkedIn URL column wider
        sheet.column_dimensions[get_column_letter(len(headers))].width = 40

        # Freeze header row
        sheet.freeze_panes = "A2"

        # Save file
        filepath = os.path.join(self.output_directory, filename)
        workbook.save(filepath)

        return filepath

    def export_summary(
        self,
        total_candidates: int,
        filtered_candidates: int,
        filter_details: dict,
        filename: str = "filter_summary.xlsx"
    ) -> str:
        """
        Export filtering summary report.

        Args:
            total_candidates: Total number of candidates before filtering
            filtered_candidates: Number of candidates after filtering
            filter_details: Detailed filter results
            filename: Output filename

        Returns:
            Full path to the exported file
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Filter Summary"

        # Title
        sheet.cell(row=1, column=1, value="Candidate Filtering Summary")
        sheet.cell(row=1, column=1).font = Font(bold=True, size=14)

        # Summary info
        sheet.cell(row=3, column=1, value="Generated:")
        sheet.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        sheet.cell(row=4, column=1, value="Total Candidates:")
        sheet.cell(row=4, column=2, value=total_candidates)

        sheet.cell(row=5, column=1, value="Passed Filters:")
        sheet.cell(row=5, column=2, value=filtered_candidates)

        sheet.cell(row=6, column=1, value="Pass Rate:")
        pass_rate = (filtered_candidates / total_candidates * 100) if total_candidates > 0 else 0
        sheet.cell(row=6, column=2, value=f"{pass_rate:.1f}%")

        # Filter breakdown
        sheet.cell(row=8, column=1, value="Filter Breakdown")
        sheet.cell(row=8, column=1).font = Font(bold=True)

        row = 9
        for filter_name, failed_count in filter_details.items():
            if filter_name.startswith("failed_"):
                display_name = filter_name.replace("failed_", "").replace("_", " ").title()
                sheet.cell(row=row, column=1, value=f"Failed {display_name}:")
                sheet.cell(row=row, column=2, value=len(failed_count) if isinstance(failed_count, list) else failed_count)
                row += 1

        # Adjust column widths
        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 25

        # Save file
        filepath = os.path.join(self.output_directory, filename)
        workbook.save(filepath)

        return filepath
